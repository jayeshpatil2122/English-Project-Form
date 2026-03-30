from copy import copy
import os
import re

from flask import Flask, redirect, render_template, request, send_file, session, url_for
from flask_sqlalchemy import SQLAlchemy
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table
from sqlalchemy import inspect, text

app = Flask(__name__)

# ===============================
# DATABASE CONFIG
# ===============================
database_url = os.environ.get("DATABASE_URL")

if database_url:
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql+psycopg://", 1)
    elif database_url.startswith("postgresql://"):
        database_url = database_url.replace("postgresql://", "postgresql+psycopg://", 1)

    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
    print("Using PostgreSQL database")
else:
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///groups.db"
    print("Using SQLite database")

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# If SECRET_KEY is missing or blank, Flask sessions break and admin login returns 500.
secret_key = (os.environ.get("SECRET_KEY") or "").strip()
app.config["SECRET_KEY"] = secret_key or "group-project-dev-secret"
db = SQLAlchemy(app)

ADMIN_PASSWORD = "1353"
MAX_GROUPS_PER_DIVISION = 26

DEFAULT_SUBJECT_KEY = "strategic-communication-for-professionals"
DEFAULT_DIVISION = "A"
ALL_DIVISIONS_KEY = "all"
DIVISIONS = ["A", "B", "C", "D", "E"]
SESSION_STUDENT_DIVISION_KEY = "student_division"
SESSION_STUDENT_PRN_KEY = "student_verified_prn"

# ===============================
# DIVISION-WISE PRN WHITELIST
# Update these lists with valid PRNs for each division.
# ===============================
DIVISION_PRN_WHITELIST = {
    "A": [
        "230105131259",
"230105131337",
"230105131347",
"230105131478",
"240101051007",
"240101081015",
"240101431007",
"240101461025",
"240102161009",
"240105011044",
"240105121007",
"240105131001",
"240105131003",
"240105131004",
"240105131005",
"240105131006",
"240105131007",
"240105131010",
"240105131011",
"240105131012",
"240105131013",
"240105131015",
"240105131016",
"240105131017",
"240105131018",
"240105131020",
"240105131021",
"240105131022",
"240105131024",
"240105131026",
"240105131027",
"240105131028",
"240105131029",
"240105131030",
"240105131031",
"240105131032",
"240105131033",
"240105131034",
"240105131035",
"240105131036",
"240105131038",
"240105131039",
"240105131041",
"240105131043",
"240105131045",
"240105131047",
"240105131048",
"240105131049",
"240105131050",
"240105131051",
"240105131053",
"240105131054",
"240105131055",
"240105131056",
"240105131058",
"240105131060",
"240105131061",
"240105131062",
"240105131063",
"240105131064",
"240105131065",
"240105131066",
"240105131068",
"240105131069",
"240105131070",
"240105131071",
"240105131072",
"240105131073",
"240105131074",
"240105131075",
"240105131076",
"240105131077",
"240105131079",
"240105131080",
"240105131081",
"240105131082",
"240105131083",
"240105131085",
"240105131086",
"240105131087",
"240105131089",
"240105131090",
"240105131091",
"240105131092",
"240105131093",
"240105131095",
"240105131096",
"240105131100",
"240105131101",
"240105131103",
"240105131104",
"240105131105",
"240105131106",
"240105131110",
"240105131111",
"240105131112",
"240105131113",
"240105131114",
"240105131116",
"240105131118",
    ],
    "B": [
        "240105131119",
"240105131120",
"240105131122",
"240105131125",
"240105131126",
"240105131130",
"240105131132",
"240105131133",
"240105131134",
"240105131135",
"240105131136",
"240105131138",
"240105131140",
"240105131141",
"240105131142",
"240105131143",
"240105131144",
"240105131146",
"240105131147",
"240105131148",
"240105131149",
"240105131150",
"240105131152",
"240105131153",
"240105131156",
"240105131157",
"240105131159",
"240105131162",
"240105131164",
"240105131165",
"240105131168",
"240105131169",
"240105131171",
"240105131172",
"240105131173",
"240105131174",
"240105131175",
"240105131176",
"240105131177",
"240105131179",
"240105131180",
"240105131182",
"240105131184",
"240105131185",
"240105131186",
"240105131188",
"240105131191",
"240105131192",
"240105131194",
"240105131195",
"240105131196",
"240105131197",
"240105131198",
"240105131199",
"240105131201",
"240105131202",
"240105131205",
"240105131206",
"240105131207",
"240105131208",
"240105131209",
"240105131210",
"240105131211",
"240105131212",
"240105131213",
"240105131214",
"240105131215",
"240105131216",
"240105131217",
"240105131218",
"240105131220",
"240105131221",
"240105131224",
"240105131226",
"240105131228",
"240105131229",
"240105131230",
"240105131231",
"240105131232",
"240105131234",
"240105131235",
"240105131236",
"240105131237",
"240105131238",
"240105131239",
"240105131241",
"240105131244",
"240105131245",
"240105131247",
"240105131248",
"240105131250",
"240105131252",
"240105131254",
"240105131255",
"240105131256",
"240105131257",
"240105131258",
"240105131263",
"240105131264",
"240105131266",
    ],
    "C": [
        "240105131267",
"240105131268",
"240105131269",
"240105131271",
"240105131272",
"240105131273",
"240105131274",
"240105131275",
"240105131276",
"240105131277",
"240105131278",
"240105131279",
"240105131280",
"240105131282",
"240105131283",
"240105131284",
"240105131285",
"240105131286",
"240105131288",
"240105131292",
"240105131293",
"240105131294",
"240105131295",
"240105131296",
"240105131298",
"240105131300",
"240105131301",
"240105131302",
"240105131303",
"240105131304",
"240105131306",
"240105131307",
"240105131309",
"240105131310",
"240105131311",
"240105131313",
"240105131315",
"240105131316",
"240105131317",
"240105131318",
"240105131319",
"240105131320",
"240105131321",
"240105131323",
"240105131324",
"240105131325",
"240105131326",
"240105131327",
"240105131328",
"240105131329",
"240105131330",
"240105131331",
"240105131332",
"240105131333",
"240105131334",
"240105131335",
"240105131336",
"240105131337",
"240105131338",
"240105131339",
"240105131340",
"240105131341",
"240105131342",
"240105131343",
"240105131344",
"240105131345",
"240105131346",
"240105131347",
"240105131348",
"240105131349",
"240105131350",
"240105131351",
"240105131353",
"240105131354",
"240105131355",
"240105131356",
"240105131357",
"240105131358",
"240105131359",
"240105131360",
"240105131362",
"240105131363",
"240105131364",
"240105131365",
"240105131366",
"240105131367",
"240105131368",
"240105131369",
"240105131370",
"240105131371",
"240105131373",
"240105131374",
"240105131375",
"240105131376",
"240105131377",
"240105131378",
"240105131379",
"240105131380",
"240105131381",
"240105131382",
    ],
    "D": [
        "240105131383",
"240105131384",
"240105131385",
"240105131386",
"240105131387",
"240105131389",
"240105131390",
"240105131391",
"240105131392",
"240105131393",
"240105131394",
"240105131395",
"240105131396",
"240105131397",
"240105131398",
"240105131399",
"240105131400",
"240105131401",
"240105131402",
"240105131403",
"240105131404",
"240105131405",
"240105131407",
"240105131408",
"240105131409",
"240105131410",
"240105131411",
"240105131412",
"240105131413",
"240105131414",
"240105131415",
"240105131416",
"240105131417",
"240105131418",
"240105131419",
"240105131420",
"240105131421",
"240105131422",
"240105131424",
"240105131425",
"240105131427",
"240105131428",
"240105131429",
"240105131430",
"240105131431",
"240105131433",
"240105131434",
"240105131435",
"240105131436",
"240105131438",
"240105131439",
"240105131442",
"240105131443",
"240105131445",
"240105131446",
"240105132024",
"240105231056",
"240105231086",
"240105231098",
"240105231135",
"240105231188",
"240105231230",
"240105231297",
"240105231382",
"240105231581",
"240105231586",
"240105231758",
"250105132001",
"250105132002",
"250105132003",
"250105132004",
"250105132005",
"250105132006",
"250105132007",
"250105132008",
"250105132009",
"250105132010",
"250105132011",
"250105132012",
"250105132013",
"250105132015",
"250105132016",
"250105132017",
"250105132019",
"250105132020",
"250105132021",
"250105132022",
"250105132023",
"250105132024",
"250105132025",
"250105132026",
"250105132027",
"250105132028",
"250105132030",
"250105132031",
"250105132032",
"250105132033",
"250105132034",
"250105132035",
"250105132036",
    ],
    "E": [
        "250105132037",
"250105132038",
"250105132039",
"250105132040",
"250105132041",
"250105132042",
"250105132043",
"250105132044",
"250105132045",
"250105132046",
"250105132048",
"250105132049",
"250105132050",
"250105132051",
"250105132052",
"250105132053",
"250105132054",
"250105132055",
"250105132056",
"250105132057",
"250105132058",
"250105132059",
"250105132060",
"250105132061",
"250105132062",
"250105132063",
"250105132064",
"250105132065",
"250105132066",
"250105132067",
"250105132068",
"250105132069",
"250105132070",
"250105132071",
"250105132072",
"250105132073",
"250105132074",
"250105132075",
"250105132076",
"250105132077",
"250105132078",
"250105132079",
"250105132080",
"250105132081",
"250105132082",
"250105132083",
"250105132084",
"250105132085",
"250105132086",
"250105132087",
"250105132088",
"250105132089",
"250105132090",
"250105132091",
"250105132092",
"250105132093",
"250105132094",
"250105132095",
"250105132096",
"250105132097",
"250105132098",
"250105132099",
"250105132100",
"250105132101",
"250105132102",
"250105132103",
"250105132105",
"250105132106",
"250105132107",
"250105132108",
"250105132109",
"250105132110",
"250105132111",
"250105132112",
"250105132113",
"250105132114",
"250105132115",
"250105132116",
"250105132117",
"250105132118",
"250105132119",
"250105132120",
"250105132122",
"250105132123",
"250105132124",
"250105132125",
"250105132126",
"250105132127",
"250105132128",
"250105132129",
"250105132130",
"250105132131",
"250105132132",
"250105132133",
"250105132134",
"250105132135",
"250105132136",
"250105133001",
    ],
}

SUBJECTS = [
    {
        #timer update nd teacher details 
        "key": DEFAULT_SUBJECT_KEY,
        "name": "Strategic Communication for Professionals",
        "faculty": "Dr.Rani Sarode",
        "deadline": "April 2, 2026 23:59:59",
        "topics": [
            {
                "title": "Organize a Mini Group Discussion",
                "description": (
                    "Students conduct a GD on a trending topic, record the session, "
                    "and prepare a self-evaluation report."
                ),
            },
            {
                "title": "Panel Discussion Simulation",
                "description": (
                    "Students form a panel with defined roles (moderator, panellists) "
                    "and conduct a formal discussion."
                ),
            },
            {
                "title": "GD Technique Handbook",
                "description": (
                    "Create a handbook of '10 Best GD Strategies' with examples and "
                    "illustrations."
                ),
            },
            {
                "title": "Active Listening Analysis",
                "description": (
                    "Observe a live/recorded debate or interview and identify examples "
                    "of constructive listening and poor listening."
                ),
            },
            {
                "title": "Barriers to Group Communication Project",
                "description": (
                    "Survey peers, identify communication barriers faced in teams, and "
                    "propose solutions."
                ),
            },
            {
                "title": "Persuasion Case Study",
                "description": (
                    "Choose a real case where persuasion changed an outcome and analyze "
                    "techniques used and ethics involved."
                ),
            },
            {
                "title": "Negotiation Role-Play",
                "description": (
                    "Students perform a negotiation scenario (salary negotiation, "
                    "conflict negotiation, business deal, etc.) and record it."
                ),
            },
            {
                "title": "Influence vs. Manipulation Poster",
                "description": (
                    "Create an infographic differentiating ethical influence and "
                    "manipulation with real-world examples."
                ),
            },
            {
                "title": "EI-Based Influence Journal",
                "description": (
                    "Maintain a journal documenting how emotional intelligence "
                    "influenced interactions over one week."
                ),
            },
            {
                "title": "Public Speaking for Influence",
                "description": (
                    "Deliver a 2-minute persuasive speech and track audience response."
                ),
            },
            {
                "title": "Personal Resilience Plan",
                "description": (
                    "Students design a step-by-step resilience-building plan for "
                    "personal or academic challenges."
                ),
            },
            {
                "title": "Adaptability Challenge Video",
                "description": (
                    "Try a new skill or unfamiliar task for 48-72 hours and document "
                    "the experience."
                ),
            },
            {
                "title": "Growth Mindset Reflection Booklet",
                "description": (
                    "Prepare a reflective booklet on replacing fixed-mindset beliefs "
                    "with growth-mindset approaches."
                ),
            },
            {
                "title": "Workplace Change Case Study",
                "description": (
                    "Analyze how a company successfully navigated change "
                    "(e.g., remote work shift, layoffs, digital transformation)."
                ),
            },
            {
                "title": "Resilience Interview Project",
                "description": (
                    "Interview a teacher/professional about how they overcame setbacks "
                    "and summarize insights."
                ),
            },
            {
                "title": "Ethical Dilemma Analysis",
                "description": (
                    "Present a real/fictional workplace dilemma and propose ethical "
                    "solutions using decision-making models."
                ),
            },
            {
                "title": "Code of Ethics Drafting Project",
                "description": (
                    "Develop a simple code of ethics for a student club, department, "
                    "or startup."
                ),
            },
            {
                "title": "Case Study: Breaking Ethical Norms",
                "description": (
                    "Analyze a corporate ethics scandal (Enron, Volkswagen, Satyam, "
                    "etc.) and identify ethical failures."
                ),
            },
            {
                "title": "Integrity in Daily Life Diary",
                "description": (
                    "Maintain a 5-day diary identifying how integrity influenced "
                    "everyday decisions."
                ),
            },
            {
                "title": "Ethics Debate Project",
                "description": (
                    "Conduct a classroom debate on topics such as 'Profit vs Ethics' "
                    "or 'Whistleblowing: Duty or Betrayal?'."
                ),
            },
            {
                "title": "Professional Etiquette Manual",
                "description": (
                    "Create a manual covering workplace dress code, email etiquette, "
                    "meeting etiquette, etc."
                ),
            },
            {
                "title": "Mock Meeting Simulation",
                "description": (
                    "Conduct a formal meeting simulation with agenda, minutes, roles, "
                    "and etiquette assessment."
                ),
            },
            {
                "title": "Office Environment Observation Project",
                "description": (
                    "Observe etiquette in a real or virtual office setting and report "
                    "professional behaviors and violations."
                ),
            },
            {
                "title": "Career Portfolio and LinkedIn Optimization",
                "description": (
                    "Build a professional digital portfolio including resume, bio, "
                    "achievements, and LinkedIn improvements."
                ),
            },
            {
                "title": "Team Culture Building Activity",
                "description": (
                    "Design a short activity or initiative to improve workplace/team "
                    "culture (icebreaker, motivation board, etc.)."
                ),
            },
        ],
    }
]
SUBJECTS_BY_KEY = {subject["key"]: subject for subject in SUBJECTS}


# ===============================
# MODEL
# ===============================
class Group(db.Model):
    __tablename__ = "groups"

    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(100), nullable=False, default=DEFAULT_SUBJECT_KEY, index=True)
    division = db.Column(db.String(1), nullable=False, default=DEFAULT_DIVISION, index=True)
    topic = db.Column(db.String(200))

    m1_name = db.Column(db.String(100))
    m1_prn = db.Column(db.String(100))
    m2_name = db.Column(db.String(100))
    m2_prn = db.Column(db.String(100))
    m3_name = db.Column(db.String(100))
    m3_prn = db.Column(db.String(100))
    m4_name = db.Column(db.String(100))
    m4_prn = db.Column(db.String(100))


class DivisionAccess(db.Model):
    __tablename__ = "division_access"

    division = db.Column(db.String(1), primary_key=True)
    is_open = db.Column(db.Boolean, nullable=False, default=True)


def ensure_subject_column():
    columns = [col["name"] for col in inspect(db.engine).get_columns("groups")]
    with db.engine.begin() as connection:
        if "subject" not in columns:
            connection.execute(text("ALTER TABLE groups ADD COLUMN subject VARCHAR(100)"))
        connection.execute(
            text(
                "UPDATE groups SET subject = :default_subject "
                "WHERE subject IS NULL OR subject = ''"
            ),
            {"default_subject": DEFAULT_SUBJECT_KEY},
        )
        connection.execute(text("CREATE INDEX IF NOT EXISTS ix_groups_subject ON groups (subject)"))


def ensure_division_column():
    columns = [col["name"] for col in inspect(db.engine).get_columns("groups")]
    with db.engine.begin() as connection:
        if "division" not in columns:
            connection.execute(text("ALTER TABLE groups ADD COLUMN division VARCHAR(1)"))
        connection.execute(
            text(
                "UPDATE groups SET division = :default_division "
                "WHERE division IS NULL OR TRIM(division) = ''"
            ),
            {"default_division": DEFAULT_DIVISION},
        )
        connection.execute(text("CREATE INDEX IF NOT EXISTS ix_groups_division ON groups (division)"))


def migrate_sqlite_drop_topic_unique(connection):
    connection.execute(
        text(
            """
            CREATE TABLE groups_new (
                id INTEGER PRIMARY KEY,
                subject VARCHAR(100) NOT NULL,
                division VARCHAR(1) NOT NULL,
                topic VARCHAR(200),
                m1_name VARCHAR(100),
                m1_prn VARCHAR(100),
                m2_name VARCHAR(100),
                m2_prn VARCHAR(100),
                m3_name VARCHAR(100),
                m3_prn VARCHAR(100),
                m4_name VARCHAR(100),
                m4_prn VARCHAR(100)
            )
            """
        )
    )
    connection.execute(
        text(
            """
            INSERT INTO groups_new (
                id, subject, division, topic, m1_name, m1_prn, m2_name, m2_prn, m3_name, m3_prn, m4_name, m4_prn
            )
            SELECT
                id,
                COALESCE(NULLIF(subject, ''), :default_subject),
                COALESCE(NULLIF(division, ''), :default_division),
                topic, m1_name, m1_prn, m2_name, m2_prn, m3_name, m3_prn, m4_name, m4_prn
            FROM groups
            """
        ),
        {"default_subject": DEFAULT_SUBJECT_KEY, "default_division": DEFAULT_DIVISION},
    )
    connection.execute(text("DROP TABLE groups"))
    connection.execute(text("ALTER TABLE groups_new RENAME TO groups"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_groups_subject ON groups (subject)"))
    connection.execute(text("CREATE INDEX IF NOT EXISTS ix_groups_division ON groups (division)"))


def ensure_topic_is_not_unique():
    unique_constraints = inspect(db.engine).get_unique_constraints("groups")
    topic_constraints = [
        constraint for constraint in unique_constraints if constraint.get("column_names") == ["topic"]
    ]
    if not topic_constraints:
        return

    with db.engine.begin() as connection:
        dialect = db.engine.dialect.name

        if dialect == "sqlite":
            migrate_sqlite_drop_topic_unique(connection)
            return

        for constraint in topic_constraints:
            constraint_name = constraint.get("name")
            if not constraint_name:
                continue
            if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", constraint_name):
                continue
            connection.execute(text(f"ALTER TABLE groups DROP CONSTRAINT {constraint_name}"))


def ensure_division_access_rows():
    existing_divisions = {row.division for row in DivisionAccess.query.all()}
    missing_divisions = [division for division in DIVISIONS if division not in existing_divisions]
    if not missing_divisions:
        return

    for division in missing_divisions:
        db.session.add(DivisionAccess(division=division, is_open=True))
    db.session.commit()


def remove_legacy_subject_data():
    with db.engine.begin() as connection:
        connection.execute(
            text(
                "DELETE FROM groups "
                "WHERE subject IS NOT NULL AND TRIM(subject) <> :default_subject"
            ),
            {"default_subject": DEFAULT_SUBJECT_KEY},
        )


with app.app_context():
    db.create_all()
    ensure_subject_column()
    remove_legacy_subject_data()
    ensure_division_column()
    ensure_topic_is_not_unique()
    ensure_division_access_rows()


# ===============================
# HELPER FUNCTIONS
# ===============================
def clean_text(value):
    return re.sub(r"\s+", "", (value or "").lower())


def build_topic_options(topics, submitted_topics):
    options = []
    for topic in topics:
        if isinstance(topic, dict):
            title = str(topic.get("title") or "").strip()
            description = str(topic.get("description") or "").strip()
            value = str(topic.get("value") or title).strip()
        else:
            value = str(topic or "").strip()
            title = value
            description = ""

        if not value:
            continue
        if not title:
            title = value

        topic_key = clean_text(value)
        options.append(
            {
                "title": title,
                "description": description,
                "value": value,
                "is_submitted": topic_key in submitted_topics,
            }
        )
    return options


def clean_words(value):
    lowered = (value or "").lower()
    lowered = re.sub(r"[^a-z\s]", "", lowered)
    return set(lowered.split())


def topics_similar(topic1, topic2):
    words1 = clean_words(topic1)
    words2 = clean_words(topic2)

    if not words1 or not words2:
        return False

    common = words1.intersection(words2)
    similarity_ratio = len(common) / min(len(words1), len(words2))
    return similarity_ratio >= 0.7


def normalize_subject_key(subject_key):
    key = (subject_key or "").strip().lower()
    if key in SUBJECTS_BY_KEY:
        return key
    return DEFAULT_SUBJECT_KEY


def get_selected_subject_key():
    return DEFAULT_SUBJECT_KEY


def normalize_division(division, allow_all=False):
    value = (division or "").strip().upper()
    if allow_all and value == ALL_DIVISIONS_KEY.upper():
        return ALL_DIVISIONS_KEY
    if value in DIVISIONS:
        return value
    return DEFAULT_DIVISION


def get_selected_division(allow_all=False):
    raw_division = request.values.get("division")
    return normalize_division(raw_division, allow_all=allow_all)


def normalize_specific_division(division):
    value = (division or "").strip().upper()
    if value in DIVISIONS:
        return value
    return None


def normalize_prn(prn):
    return re.sub(r"\D", "", str(prn or ""))


def is_valid_prn_format(prn):
    normalized = normalize_prn(prn)
    return normalized.isdigit() and len(normalized) == 12


def get_allowed_prns_for_division(division):
    if division not in DIVISIONS:
        return set()
    return {
        normalize_prn(item)
        for item in DIVISION_PRN_WHITELIST.get(division, [])
        if is_valid_prn_format(item)
    }


def is_prn_allowed_for_division(division, prn):
    normalized_division = normalize_specific_division(division)
    normalized_prn = normalize_prn(prn)
    if not normalized_division or not is_valid_prn_format(normalized_prn):
        return False
    return normalized_prn in get_allowed_prns_for_division(normalized_division)


def clear_student_access_session():
    session.pop(SESSION_STUDENT_DIVISION_KEY, None)
    session.pop(SESSION_STUDENT_PRN_KEY, None)


def get_verified_student_access():
    division = normalize_specific_division(session.get(SESSION_STUDENT_DIVISION_KEY))
    prn = normalize_prn(session.get(SESSION_STUDENT_PRN_KEY))
    if not division or not is_prn_allowed_for_division(division, prn):
        clear_student_access_session()
        return None, None
    return division, prn


def get_groups_for_scope(subject_key, division):
    query = Group.query.filter_by(subject=subject_key)
    if division != ALL_DIVISIONS_KEY:
        query = query.filter_by(division=division)
    return query.order_by(Group.id.asc()).all()


def get_division_access_map():
    division_access_map = {division: True for division in DIVISIONS}
    for row in DivisionAccess.query.all():
        if row.division in division_access_map:
            division_access_map[row.division] = bool(row.is_open)
    return division_access_map


def admin_required_redirect():
    if not session.get("is_admin"):
        return redirect(url_for("admin"))
    return None


# ===============================
# STUDENT PAGE
# ===============================
@app.route("/", methods=["GET", "POST"])
def index():
    popup = None
    message = None
    access_message = None

    selected_subject_key = get_selected_subject_key()
    selected_subject = SUBJECTS_BY_KEY[selected_subject_key]

    selected_division, _verified_prn = get_verified_student_access()
    division_access_map = get_division_access_map()
    all_topics = selected_subject["topics"]

    def render_gate(prefill_division=""):
        return render_template(
            "index.html",
            groups=[],
            popup=popup,
            message=message,
            access_message=access_message,
            show_access_gate=True,
            access_division=prefill_division,
            topic_options=build_topic_options(all_topics, set()),
            submitted_topics=set(),
            selected_subject=selected_subject,
            selected_subject_key=selected_subject_key,
            selected_division=selected_division,
            divisions=DIVISIONS,
            max_groups_per_division=MAX_GROUPS_PER_DIVISION,
            selected_division_open=False,
            division_access_map=division_access_map,
        )

    if request.method == "POST" and request.form.get("action") == "verify_prn":
        division_input = normalize_specific_division(request.form.get("division"))
        prn_input = normalize_prn(request.form.get("access_prn"))

        if not division_input:
            access_message = "Please enter a valid division (A/B/C/D/E)."
            return render_gate()

        if not is_valid_prn_format(prn_input):
            access_message = "PRN must be exactly 12 digits."
            return render_gate(prefill_division=division_input)

        if not is_prn_allowed_for_division(division_input, prn_input):
            access_message = f"Access denied: PRN is not in Div {division_input}."
            return render_gate(prefill_division=division_input)

        session[SESSION_STUDENT_DIVISION_KEY] = division_input
        session[SESSION_STUDENT_PRN_KEY] = prn_input
        return redirect(url_for("index"))

    if not selected_division:
        return render_gate()

    all_groups_for_division = get_groups_for_scope(selected_subject_key, selected_division)
    selected_division_open = division_access_map.get(selected_division, True)
    existing_groups = all_groups_for_division if selected_division_open else []
    submitted_topics = {clean_text(group.topic) for group in existing_groups if group.topic}

    def render_index():
        return render_template(
            "index.html",
            groups=existing_groups,
            popup=popup,
            message=message,
            access_message=access_message,
            show_access_gate=False,
            access_division=selected_division,
            topic_options=build_topic_options(all_topics, submitted_topics),
            submitted_topics=submitted_topics,
            selected_subject=selected_subject,
            selected_subject_key=selected_subject_key,
            selected_division=selected_division,
            divisions=DIVISIONS,
            max_groups_per_division=MAX_GROUPS_PER_DIVISION,
            selected_division_open=selected_division_open,
            division_access_map=division_access_map,
        )

    if request.method == "POST":
        if not selected_division_open:
            popup = "division_closed"
            message = f"Closed: Data is not visible because Div {selected_division} form is closed."
            return render_index()

        if (
            Group.query.filter_by(subject=selected_subject_key, division=selected_division).count()
            >= MAX_GROUPS_PER_DIVISION
        ):
            popup = "max_groups"
            message = f"Maximum {MAX_GROUPS_PER_DIVISION} groups allowed for Div {selected_division}."
            return render_index()

        topic = request.form.get("topic", "").strip()
        if not topic:
            popup = "invalid_topic"
            message = "Topic is required."
            return render_index()

        for group in existing_groups:
            if topics_similar(topic, group.topic):
                popup = "duplicate_topic"
                message = (
                    f"Topic already selected by Group #{group.id} in Div {selected_division}."
                )
                return render_index()

        members = []
        for index in range(1, 5):
            name = request.form.get(f"m{index}_name", "").strip()
            prn = request.form.get(f"m{index}_prn", "").strip()
            if name and prn:
                members.append((name, normalize_prn(prn)))

        if len(members) == 0:
            popup = "invalid_group"
            message = "At least 1 member is required."
            return render_index()

        for _name, prn in members:
            if not is_valid_prn_format(prn):
                popup = "invalid_prn"
                message = f"PRN {prn} must be exactly 12 digits."
                return render_index()
            if not is_prn_allowed_for_division(selected_division, prn):
                popup = "invalid_prn"
                message = f"PRN {prn} is not in Div {selected_division} list."
                return render_index()

        for group in existing_groups:
            existing_names = [group.m1_name, group.m2_name, group.m3_name, group.m4_name]
            existing_prns = [group.m1_prn, group.m2_prn, group.m3_prn, group.m4_prn]

            for name, prn in members:
                if clean_text(prn) in [clean_text(existing) for existing in existing_prns if existing]:
                    popup = "duplicate_user"
                    message = (
                        f"PRN {prn} already in Group #{group.id} for Div {selected_division}."
                    )
                    return render_index()

                if clean_text(name) in [clean_text(existing) for existing in existing_names if existing]:
                    popup = "duplicate_user"
                    message = (
                        f"{name} already in Group #{group.id} for Div {selected_division}."
                    )
                    return render_index()

        new_group = Group(topic=topic, subject=selected_subject_key, division=selected_division)

        if len(members) >= 1:
            new_group.m1_name, new_group.m1_prn = members[0]
        if len(members) >= 2:
            new_group.m2_name, new_group.m2_prn = members[1]
        if len(members) >= 3:
            new_group.m3_name, new_group.m3_prn = members[2]
        if len(members) >= 4:
            new_group.m4_name, new_group.m4_prn = members[3]

        db.session.add(new_group)
        db.session.commit()

        return redirect(url_for("index"))

    return render_index()


@app.route("/student/logout", methods=["POST"])
def student_logout():
    clear_student_access_session()
    return redirect(url_for("index"))


# ===============================
# ADMIN LOGIN + PANEL
# ===============================
@app.route("/admin", methods=["GET", "POST"])
def admin():
    if request.method == "POST" and "password" in request.form:
        if request.form["password"] == ADMIN_PASSWORD:
            session["is_admin"] = True
            return redirect(url_for("admin", division=get_selected_division(allow_all=True)))
        return render_template("admin_login.html", error="Wrong Password!")

    if not session.get("is_admin"):
        return render_template("admin_login.html")

    if request.method == "POST" and request.form.get("action") == "set_division_access":
        division = normalize_division(request.form.get("division"))
        is_open = request.form.get("is_open") == "1"

        division_access = db.session.get(DivisionAccess, division)
        if division_access is None:
            division_access = DivisionAccess(division=division, is_open=is_open)
            db.session.add(division_access)
        else:
            division_access.is_open = is_open
        db.session.commit()

        return redirect(url_for("admin", division=division))

    selected_subject_key = get_selected_subject_key()
    selected_subject = SUBJECTS_BY_KEY[selected_subject_key]
    selected_division = get_selected_division(allow_all=True)

    groups = get_groups_for_scope(selected_subject_key, selected_division)
    division_access_map = get_division_access_map()
    selected_division_is_specific = selected_division != ALL_DIVISIONS_KEY
    division_is_open = (
        division_access_map.get(selected_division, True) if selected_division_is_specific else None
    )

    return render_template(
        "admin.html",
        groups=groups,
        selected_subject=selected_subject,
        selected_division=selected_division,
        selected_division_is_specific=selected_division_is_specific,
        divisions=DIVISIONS,
        all_divisions_key=ALL_DIVISIONS_KEY,
        division_access_map=division_access_map,
        division_is_open=division_is_open,
    )


@app.route("/admin/logout", methods=["GET", "POST"])
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin"))


# ===============================
# ADMIN CRUD
# ===============================
@app.route("/admin/edit/<int:group_id>", methods=["GET", "POST"])
def edit_group(group_id):
    admin_redirect = admin_required_redirect()
    if admin_redirect:
        return admin_redirect

    group = Group.query.get_or_404(group_id)
    selected_division_view = normalize_division(
        request.values.get("selected_division"), allow_all=True
    )

    if request.method == "POST":
        selected_division = normalize_division(request.form.get("division") or group.division)
        topic = request.form.get("topic", "").strip()

        if topic:
            group.topic = topic

        group.subject = DEFAULT_SUBJECT_KEY
        group.division = selected_division

        for index in range(1, 5):
            name = request.form.get(f"m{index}_name", "").strip()
            prn = request.form.get(f"m{index}_prn", "").strip()
            setattr(group, f"m{index}_name", name or None)
            setattr(group, f"m{index}_prn", prn or None)

        db.session.commit()
        return redirect(url_for("admin", division=selected_division_view))

    selected_division = normalize_division(request.args.get("division") or group.division)
    return render_template(
        "edit_group.html",
        group=group,
        divisions=DIVISIONS,
        selected_subject=SUBJECTS_BY_KEY[DEFAULT_SUBJECT_KEY],
        selected_division=selected_division,
        selected_division_view=selected_division_view,
    )


@app.route("/admin/delete/<int:group_id>", methods=["POST"])
def delete_group(group_id):
    admin_redirect = admin_required_redirect()
    if admin_redirect:
        return admin_redirect

    group = Group.query.get_or_404(group_id)
    selected_division_view = normalize_division(
        request.form.get("selected_division"), allow_all=True
    )
    db.session.delete(group)
    db.session.commit()
    return redirect(url_for("admin", division=selected_division_view))


# ===============================
# DOWNLOAD EXCEL (CIA FORMAT STYLE)
# ===============================
@app.route("/download_excel")
def download_excel():
    admin_redirect = admin_required_redirect()
    if admin_redirect:
        return admin_redirect

    from openpyxl import Workbook, load_workbook

    selected_subject_key = get_selected_subject_key()
    selected_subject = SUBJECTS_BY_KEY[selected_subject_key]
    selected_division = get_selected_division(allow_all=True)
    groups = get_groups_for_scope(selected_subject_key, selected_division)

    include_division_in_topic = selected_division == ALL_DIVISIONS_KEY
    division_label = (
        "All Divisions" if include_division_in_topic else f"Div {selected_division}"
    )

    template_candidates = [
        os.environ.get("EXCEL_TEMPLATE_PATH"),
        os.path.join(app.root_path, "Strategic Communication CIA 3 Updated.xlsx"),
    ]
    template_path = next((path for path in template_candidates if path and os.path.exists(path)), None)

    def clone_row_style(sheet, src_row, dst_row, max_col=4):
        src_dim = sheet.row_dimensions[src_row]
        if src_dim.height is not None:
            sheet.row_dimensions[dst_row].height = src_dim.height
        for column in range(1, max_col + 1):
            src = sheet.cell(src_row, column)
            dst = sheet.cell(dst_row, column)
            if src.has_style:
                dst._style = copy(src._style)

    def next_serial_number(sheet):
        serials = []
        for row in range(5, sheet.max_row + 1):
            value = sheet.cell(row, 1).value
            if value in (None, ""):
                continue
            try:
                serials.append(int(str(value).strip()))
            except ValueError:
                continue
        return (max(serials) + 1) if serials else 1

    def topic_for_export(group):
        topic_text = group.topic or ""
        if include_division_in_topic:
            return f"[Div {group.division}] {topic_text}".strip()
        return topic_text

    def append_groups_like_template(sheet, start_row, groups_to_write, start_serial):
        row = start_row
        serial_no = start_serial
        style_rows = [5, 6, 7, 8] if sheet.max_row >= 8 else []

        for group in groups_to_write:
            members = []
            for name, prn in [
                (group.m1_name, group.m1_prn),
                (group.m2_name, group.m2_prn),
                (group.m3_name, group.m3_prn),
                (group.m4_name, group.m4_prn),
            ]:
                if name or prn:
                    members.append((name or "", prn or ""))

            if not members:
                members = [("", "")]

            for index in range(len(members)):
                if style_rows:
                    clone_row_style(sheet, style_rows[min(index, len(style_rows) - 1)], row + index)

            if len(members) > 1:
                sheet.merge_cells(
                    start_row=row,
                    start_column=1,
                    end_row=row + len(members) - 1,
                    end_column=1,
                )
                sheet.merge_cells(
                    start_row=row,
                    start_column=4,
                    end_row=row + len(members) - 1,
                    end_column=4,
                )

            sheet.cell(row=row, column=1).value = serial_no
            sheet.cell(row=row, column=4).value = topic_for_export(group)

            for index, (name, prn) in enumerate(members):
                current_row = row + index
                sheet.cell(row=current_row, column=2).value = str(prn) if prn else ""
                sheet.cell(row=current_row, column=3).value = name

            row += len(members)
            serial_no += 1

    if template_path:
        workbook = load_workbook(template_path)
        worksheet = workbook.active

        if worksheet["A2"].value:
            worksheet["A2"] = f"Program: B.Tech CSE | Sem IV | {division_label}"
        if worksheet["A3"].value:
            worksheet["A3"] = f"CIA 3 {selected_subject['name']} list"

        start_row = worksheet.max_row + 1
        append_groups_like_template(
            worksheet,
            start_row,
            groups,
            start_serial=next_serial_number(worksheet),
        )
    else:
        from openpyxl.styles import Alignment, Border, Font, Side

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Mini Project List"

        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "Sandip University, Nashik (MS), India"
        worksheet["A1"].alignment = Alignment(horizontal="center")
        worksheet["A1"].font = Font(size=14, bold=True)

        worksheet.merge_cells("A2:D2")
        worksheet["A2"] = f"Program: B.Tech CSE | Sem IV | {division_label}"
        worksheet["A2"].alignment = Alignment(horizontal="center")

        worksheet.merge_cells("A3:D3")
        worksheet["A3"] = f"Subject: {selected_subject['name']}"
        worksheet["A3"].alignment = Alignment(horizontal="center")

        worksheet.merge_cells("A4:D4")
        worksheet["A4"] = "Mini Project List"
        worksheet["A4"].alignment = Alignment(horizontal="center")
        worksheet["A4"].font = Font(size=13, bold=True)

        worksheet["A6"] = "Sr.No"
        worksheet["B6"] = "PRN No"
        worksheet["C6"] = "Project Group Members"
        worksheet["D6"] = "Project Title"

        for column in range(1, 5):
            worksheet.cell(row=6, column=column).font = Font(bold=True)

        row = 7
        serial_no = 1
        for group in groups:
            members = []
            for name, prn in [
                (group.m1_name, group.m1_prn),
                (group.m2_name, group.m2_prn),
                (group.m3_name, group.m3_prn),
                (group.m4_name, group.m4_prn),
            ]:
                if name or prn:
                    members.append((name or "", prn or ""))

            if not members:
                members = [("", "")]

            worksheet.cell(row=row, column=1).value = serial_no
            worksheet.cell(row=row, column=4).value = topic_for_export(group)

            for index, (name, prn) in enumerate(members):
                current_row = row + index
                worksheet.cell(row=current_row, column=2).value = str(prn) if prn else ""
                worksheet.cell(row=current_row, column=3).value = name

            row += len(members)
            serial_no += 1

        worksheet.column_dimensions["A"].width = 8
        worksheet.column_dimensions["B"].width = 18
        worksheet.column_dimensions["C"].width = 25
        worksheet.column_dimensions["D"].width = 30

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for data_row in range(6, row):
            for data_col in range(1, 5):
                worksheet.cell(row=data_row, column=data_col).border = thin_border

    safe_subject = selected_subject_key.replace("-", "_")
    safe_division = (
        "all_divisions" if include_division_in_topic else f"div_{selected_division.lower()}"
    )
    file_path = f"groups_{safe_subject}_{safe_division}.xlsx"
    workbook.save(file_path)
    return send_file(
        file_path,
        as_attachment=True,
        download_name=f"{selected_subject['name']} - {division_label} Groups.xlsx",
    )


# ===============================
# DOWNLOAD PDF
# ===============================
@app.route("/download_pdf")
def download_pdf():
    admin_redirect = admin_required_redirect()
    if admin_redirect:
        return admin_redirect

    selected_subject_key = get_selected_subject_key()
    selected_subject = SUBJECTS_BY_KEY[selected_subject_key]
    selected_division = get_selected_division(allow_all=True)
    groups = get_groups_for_scope(selected_subject_key, selected_division)

    include_division_in_topic = selected_division == ALL_DIVISIONS_KEY
    division_label = (
        "All Divisions" if include_division_in_topic else f"Div {selected_division}"
    )

    safe_subject = selected_subject_key.replace("-", "_")
    safe_division = (
        "all_divisions" if include_division_in_topic else f"div_{selected_division.lower()}"
    )
    file_path = f"groups_{safe_subject}_{safe_division}.pdf"
    document = SimpleDocTemplate(file_path, pagesize=A4)

    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, Spacer, TableStyle

    elements = []
    styles = getSampleStyleSheet()

    elements.append(Spacer(1, 40))

    university_style = ParagraphStyle(
        "UniversityStyle",
        parent=styles["Title"],
        alignment=1,
        fontSize=18,
        spaceAfter=10,
    )

    normal_center = ParagraphStyle(
        "NormalCenter",
        parent=styles["Normal"],
        alignment=1,
        fontSize=12,
    )

    main_heading = ParagraphStyle(
        "MainHeading",
        parent=styles["Heading1"],
        alignment=1,
        fontSize=20,
        spaceBefore=15,
        spaceAfter=20,
    )

    elements.append(Paragraph("<b>Sandip University, Nashik (MS), India</b>", university_style))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("Program: B.Tech CSE", normal_center))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(f"Sem IV | {division_label}", normal_center))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(f"Subject: {selected_subject['name']}", normal_center))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(f"Faculty: {selected_subject['faculty']}", normal_center))
    elements.append(Spacer(1, 25))
    elements.append(Paragraph("<b>Mini Project List</b>", main_heading))
    elements.append(Spacer(1, 15))

    data = [["Sr.no", "PRN No", "Project group members", "Project title"]]

    for serial_no, group in enumerate(groups, start=1):
        prns = []
        names = []

        if group.m1_prn:
            prns.append(f"1) {group.m1_prn}")
        if group.m2_prn:
            prns.append(f"2) {group.m2_prn}")
        if group.m3_prn:
            prns.append(f"3) {group.m3_prn}")
        if group.m4_prn:
            prns.append(f"4) {group.m4_prn}")

        if group.m1_name:
            names.append(group.m1_name)
        if group.m2_name:
            names.append(group.m2_name)
        if group.m3_name:
            names.append(group.m3_name)
        if group.m4_name:
            names.append(group.m4_name)

        topic_text = group.topic or ""
        if include_division_in_topic:
            topic_text = f"[Div {group.division}] {topic_text}".strip()

        data.append([f"{serial_no})", "\n".join(prns), "\n".join(names), topic_text])

    table = Table(data, colWidths=[40, 120, 160, 150])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(table)

    document.build(elements)

    return send_file(
        file_path,
        as_attachment=True,
        download_name=f"{selected_subject['name']} - {division_label} Groups.pdf",
    )


if __name__ == "__main__":
    app.run(debug=True, port=5001)
